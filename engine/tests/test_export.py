"""Tests for mmra_engine.export."""

import os
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook

from mmra_engine.export import (
    simulation_result_to_csv_tables,
    write_simulation_result_csv,
    write_simulation_result_excel,
)
from mmra_engine.simulation import run_simulation
from mmra_engine.validation import validate_simulation_input
from mmra_engine.validation_cases import (
    PM3XD_WORKBOOK_TARGETS,
    make_pm3xd_h1ss10_input,
)


class TestExport(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.inp = make_pm3xd_h1ss10_input(seed=42, n_iterations=5000)
        cls.report = validate_simulation_input(cls.inp)
        cls.result = run_simulation(cls.inp)

    def test_csv_tables_contain_results(self):
        tables = simulation_result_to_csv_tables(
            self.result,
            input=self.inp,
            validation_report=self.report,
            workbook_targets=PM3XD_WORKBOOK_TARGETS,
        )
        self.assertIn("Metadata", tables)
        self.assertIn("Results", tables)
        self.assertIn("QAQC", tables)
        self.assertIn("MMBOE", tables["Results"] or "")

    def test_excel_sheet_names(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "out.xlsx"
            write_simulation_result_excel(
                path,
                self.result,
                input=self.inp,
                validation_report=self.report,
                workbook_targets=PM3XD_WORKBOOK_TARGETS,
            )
            wb = load_workbook(path, read_only=True)
            self.assertIn("Metadata", wb.sheetnames)
            self.assertIn("Results", wb.sheetnames)
            self.assertIn("QAQC", wb.sheetnames)
            self.assertIn("Validation_Targets", wb.sheetnames)
            wb.close()

    def test_csv_zip_export_matches_result(self):
        with tempfile.TemporaryDirectory() as tmp:
            zip_path = Path(tmp) / "export.zip"
            write_simulation_result_csv(
                zip_path,
                self.result,
                input=self.inp,
                validation_report=self.report,
            )
            with zipfile.ZipFile(zip_path) as zf:
                results_csv = zf.read("Results.csv").decode("utf-8")
            self.assertIn("MMBOE", results_csv)
            mean_str = f"{self.result.total_mmboe.mean_trimmed:.6g}"
            self.assertIn(mean_str[:6], results_csv)

    def test_metadata_percentile_convention(self):
        tables = simulation_result_to_csv_tables(self.result)
        self.assertIn(self.result.percentile_convention, tables["Metadata"])


if __name__ == "__main__":
    unittest.main()
